from openpyxl import load_workbook

ruta_archivo = "C:\\Users\\cdelgadb\\Documents\\prueba2.xlsx"

wb = load_workbook(ruta_archivo)
ws = wb.active  # O ws = wb["NombreHoja"] si necesitas una hoja específica

# Limpiar filtros existentes
ws.auto_filter.ref = None
ws.auto_filter = None
ws.sheet_view.filterMode = False

# Determinar la última fila con contenido
ultima_fila = 0
for fila in ws.iter_rows():
    if any(celda.value not in (None, "") for celda in fila):
        ultima_fila = fila[0].row

# Escribir "hola" después de esa fila en la primera columna
nueva_fila = ultima_fila + 1
ws.cell(row=nueva_fila, column=1, value="hola")

# Guardar los cambios
wb.save(ruta_archivo)
