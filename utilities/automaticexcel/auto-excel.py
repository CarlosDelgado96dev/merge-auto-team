import re
import sys
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

ruta = sys.argv[1]
resultTxt = sys.argv[2]


class Test:
    ALLOWED_SPA = ('Ficha de Cliente', 'Pangea')
    ALLOWED_ENT = ('PROD', 'ASEG', 'ENT1', 'ENT2')

    def __init__(self, spa: str, ent: str, front, message, src, date, job, step, responsibleEntity, user):
        if spa not in self.ALLOWED_SPA:
            raise ValueError(
                f"Valor inválido para spa: {spa}. "
                f"Solo se permiten: {self.ALLOWED_SPA}"
            )
        if ent not in self.ALLOWED_ENT:
            raise ValueError(
                f"Valor inválido para ent: {ent}. "
                f"Solo se permiten: {self.ALLOWED_ENT}"
            )

        self.spa = spa
        self.ent = ent
        self.front = front
        self.message = message
        self.src = src
        self.date = date
        self.job = job
        self.step = step
        self.responsibleEntity = responsibleEntity
        self.user = user

    def convertObjectForExcelLines(listObjects):
        for test in listTest:
            if 'Ficha de Cliente' in test.spa:
                test.spa = 'FDC'

            if 'Pangea' in test.spa:
                test.spa = 'PDV'

            test.src = 'JOB'
            
            # Obtener la fecha actual
            fecha_actual = datetime.now()
            test.date = fecha_actual.strftime("%d/%m/%Y")

            test.job = execution

            test.step = '0'

            test.responsibleEntity = 'Application'

            test.user = userSystemToUserInExcel(user)
            


    def toString(self):
        return f"test: {self.front} SPA: {self.spa}  entorno: {self.ent} SRC: {self.src} Date: {self.date} JOB: {self.job} Step: {self.step} Mensaje: {self.message} Rentity: {self.responsibleEntity} User: {self.user} "


def sanitize_text(value):
    if value is None:
        return value
    return ''.join(
        ch for ch in value
        if ch in ("\t", "\n", "\r") or ord(ch) >= 32
    ).strip()

def tests_to_dataframe(tests):
    registros = []
    for test in tests:
        registros.append({
            "FRONT": sanitize_text(test.front),
            "SRC": sanitize_text(test.src),
            "SPA": sanitize_text(test.spa),
            "ENT": sanitize_text(test.ent),
            "DATE": sanitize_text(test.date),
            "JOB": sanitize_text(test.job),
            "JOBS": sanitize_text(test.job),
            "STEP": sanitize_text(test.step),
            "RESPONSIBLE": sanitize_text(test.responsibleEntity),
            "ERROR TYPE": "Element not rendered",
            "ALLURE MESSAGE": sanitize_text(test.message),
            "ANALYSIS": " ",
            "STATUS": "0.Define",
            "USUARIO": sanitize_text(test.user),
        })
    return pd.DataFrame(registros)


def convertUserAndExecution(ruta):
    partes = ruta.split("/")
    user = partes[2]
    nombre_archivo = os.path.basename(ruta)
    coincidencia = re.search(r"#(\d+)", nombre_archivo)
    execution = coincidencia.group(1) if coincidencia else None
    return user, execution

user, execution = convertUserAndExecution(ruta)


print("user:", user)
print("numero_ejecucion:", execution)

listTest = []


contador = 0
for linea in resultTxt.splitlines():
    print(linea)
    spa_encontrado = None
    ent_encontrado = None
    front_code = None
    for spa in Test.ALLOWED_SPA:
        if spa in linea:
            spa_encontrado = spa
            break
    for ent in Test.ALLOWED_ENT:
        if ent in linea:
            ent_encontrado = ent
            break

    if 'FRONT' in linea:
        # Busca algo que empiece por FRONT y llegue hasta el primer punto
        match = re.search(r'(FRONT[^.]*)\.', linea)
        if match:
            front_code = match.group(1).strip()

    if 'Failed' in linea:
        match_failed = re.search(r'Failed:\s*(.+)', linea)
        if match_failed:
            message_failed = match_failed.group(1).strip()
            if listTest:
                listTest[-1].message = message_failed
        
    
    if spa_encontrado and ent_encontrado and front_code:
        listTest.append(Test(spa_encontrado, ent_encontrado, front_code, None , None , None , None , None , None, None))
        contador += 1
        
print("La lista tiene un total de elementos de " + str(len(listTest)) + " tests")

def userSystemToUserInExcel(user):
    if 'cdelgadb' in user:
        return 'Carlos Delgado Benito'
    
    if 'smonfort' in user:
        return 'Sergi Monfort Ferrer'
    
    if 'aclemens' in user:
        return 'Adrián Clement Sax'


Test.convertObjectForExcelLines(listTest)


df = tests_to_dataframe(listTest)
print(df)


ruta_excel = os.path.join(f"C:/Users/{user}/Downloads", f"resultados_#{execution}.xlsx")

df.to_excel(ruta_excel, index=False, engine="openpyxl")

workbook = load_workbook(ruta_excel)
sheet = workbook.active
alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

for row in sheet.iter_rows(
    min_row=1,
    max_row=sheet.max_row,
    min_col=1,
    max_col=sheet.max_column,
):
    for cell in row:
        cell.alignment = alignment

workbook.save(ruta_excel)

# ALINEAR EN EL CENTRO DIRECTAMENTE AL CREARLO
#with pd.ExcelWriter(ruta_excel, engine="xlsxwriter") as writer:
#    df.to_excel(writer, index=False, sheet_name="Resultados")
#    workbook = writer.book
#    worksheet = writer.sheets["Resultados"]
#    formato_centrado = workbook.add_format({"align": "center", "valign": "vcenter"})
#    worksheet.set_column(0, df.shape[1] - 1, None, formato_centrado)
#    worksheet.set_default_row(15, formato_centrado)
