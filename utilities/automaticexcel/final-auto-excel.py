import re
import sys
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font

# =========================
# 1. DETECCIÓN USUARIO Y DOWNLOADS
# =========================

def detectar_usuario_windows() -> str:
    return os.environ.get("USERNAME") or os.environ.get("USER") or ""

def detectar_downloads() -> Path:
    user = detectar_usuario_windows()
    if not user:
        raise RuntimeError("No se pudo detectar el usuario Windows")

    posibles = [
        Path(f"C:/Users/{user}/Downloads"),
        Path(f"/mnt/c/Users/{user}/Downloads"),
        Path(f"/c/Users/{user}/Downloads"),
    ]

    for p in posibles:
        if p.is_dir():
            return p

    raise RuntimeError("No se encontró la carpeta Downloads")

def buscar_archivo_por_ejecucion(downloads: Path, ejecucion: str) -> Path:
    expected_name = f"#{ejecucion}.txt"
    # Listar últimos 50 archivos por fecha de modificación
    archivos = sorted(
        downloads.glob("*"),
        key=lambda p: p.stat().st_mtime,
        reverse=True
    )[:50]

    for f in archivos:
        if f.is_file() and f.name == expected_name:
            return f

    raise FileNotFoundError(
        f"No se encontró '{expected_name}' entre los últimos 50 archivos en {downloads}"
    )


# =========================
# 2. FUNCIÓN PARA MAPEAR USER + EXECUTION DESDE LA RUTA
# =========================

def convertUserAndExecution(ruta: str):
    """
    ruta: string tipo 'C:/Users/<user>/Downloads/#1348.txt' o similar
    """
    partes = ruta.replace("\\", "/").split("/")
    # Buscar índice de 'Users' y tomar el siguiente como usuario
    user = None
    if "Users" in partes:
        idx = partes.index("Users")
        if idx + 1 < len(partes):
            user = partes[idx + 1]
    if not user and len(partes) > 2:
        # fallback muy básico
        user = partes[2]

    nombre_archivo = os.path.basename(ruta)
    coincidencia = re.search(r"#(\d+)", nombre_archivo)
    execution = coincidencia.group(1) if coincidencia else None
    return user, execution


def extraer_bloques_failures(contenido):
    """
    Extrae todos los bloques desde 'Failures' hasta 'Executed'
    y los devuelve como un único texto concatenado.
    """
    bloques = []

    # Regex: desde Failures hasta Executed (incluyéndolo)
    patron = re.compile(
        r'Failures[\s\S]*?Executed.*?$',
        re.IGNORECASE | re.MULTILINE
    )

    for match in patron.finditer(contenido):
        bloques.append(match.group())

    return "\n\n".join(bloques)

# =========================
# 3. LÓGICA DE NEGOCIO (TU CÓDIGO)
# =========================

class Test:
    ALLOWED_SPA = ('Ficha de Cliente', 'Pangea', 'Jazztel')
    ALLOWED_ENT = ('PROD', 'ASEG', 'ENT1', 'ENT2')

    def __init__(self, spa: str, ent: str, front, message, src, date, job, step, responsibleEntity, user):
        if spa not in self.ALLOWED_SPA:
            raise ValueError(
                f"Valor inválido para spa: {spa}. "
                f"Solo se permiten: {self.ALLOWED_SPA}"
            )
        if ent not in self.ALLOWED_ENT:
            raise ValueError(
                f"Valor inválido para ent: {ent}. "
                f"Solo se permiten: {self.ALLOWED_ENT}"
            )

        self.spa = spa
        self.ent = ent
        self.front = front
        self.message = message
        self.src = src
        self.date = date
        self.job = job
        self.step = step
        self.responsibleEntity = responsibleEntity
        self.user = user

    @staticmethod
    def convertObjectForExcelLines(listObjects, execution, user):
        for test in listObjects:
            if 'Ficha de Cliente' in test.spa:
                test.spa = 'FDC'

            if 'Pangea' in test.spa:
                test.spa = 'PDV'

            if 'Jazztel' in test.spa:
                test.spa = 'PDV-MM'

            test.src = 'JOB'
            
            # Fecha actual
            fecha_actual = datetime.now()
            test.date = fecha_actual.strftime("%d/%m/%Y")

            test.job = execution
            test.responsibleEntity = 'Application'
            test.user = userSystemToUserInExcel(user)

    def toString(self):
        return (
            f"test: {self.front} SPA: {self.spa}  entorno: {self.ent} "
            f"SRC: {self.src} Date: {self.date} JOB: {self.job} Step: {self.step} "
            f"Mensaje: {self.message} Rentity: {self.responsibleEntity} User: {self.user}"
        )


def sanitize_text(value):
    if value is None:
        return value
    return ''.join(
        ch for ch in value
        if ch in ("\t", "\n", "\r") or ord(ch) >= 32
    ).strip()

def tests_to_dataframe(tests):
    registros = []
    for test in tests:
        registros.append({
            "FRONT": sanitize_text(test.front),
            "SRC": sanitize_text(test.src),
            "SPA": sanitize_text(test.spa),
            "ENT": sanitize_text(test.ent),
            "DATE": sanitize_text(test.date),
            "JOB": sanitize_text(test.job),
            "JOBS": sanitize_text(test.job),
            "STEP": sanitize_text(test.step),
            "RESPONSIBLE": sanitize_text(test.responsibleEntity),
            "ERROR TYPE": "Element not rendered",
            "ALLURE MESSAGE": sanitize_text(test.message),
            "ANALYSIS": " ",
            "STATUS": "0.Define",
            "USUARIO": sanitize_text(test.user),
        })
    return pd.DataFrame(registros)

def userSystemToUserInExcel(user):
    if 'cdelgadb' in user:
        return 'Carlos Delgado Benito'
    if 'smonfort' in user:
        return 'Sergi Monfort Ferrer'
    if 'aclemens' in user:
        return 'Adrián Clement Sax'
    return user  # fallback: devolver el login si no está mapeado


# =========================
# 4. MAIN: UNE TODO
# =========================

def main():
    # 4.1 Pedir número de ejecución
    ejecucion = input("Introduce el número de ejecución (ej: 1348): ").strip()
    if not ejecucion.isdigit():
        print("Error: Debes introducir un número.", file=sys.stderr)
        sys.exit(1)

    # 4.2 Detectar Downloads
    try:
        downloads = detectar_downloads()
    except Exception as e:
        print(f"Error detectando Downloads: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"Carpeta Downloads detectada: {downloads}")

    # 4.3 Buscar archivo #<ejecucion>.txt
    try:
        ruta_archivo = buscar_archivo_por_ejecucion(downloads, ejecucion)
    except Exception as e:
        print(e, file=sys.stderr)
        sys.exit(1)

    print(f"Archivo encontrado: {ruta_archivo}")

    # 4.4 Leer contenido del archivo
    with ruta_archivo.open('r', encoding='utf-8', errors='replace') as f:
        contenido = f.read()
        contenido_failures = extraer_bloques_failures(contenido)

        if not contenido_failures:
            print("No se encontraron bloques de Failures.", file=sys.stderr)
            sys.exit(1)

        print("Bloques de Failures detectados:\n")
        print(contenido_failures)

    # 4.5 Obtener user y execution desde la ruta
    user, execution = convertUserAndExecution(str(ruta_archivo))
    if not user or not execution:
        print("No se pudo extraer user o execution desde la ruta del archivo.", file=sys.stderr)
        sys.exit(1)

    print("user:", user)
    print("numero_ejecucion:", execution)

    listTest = []
    contador = 0

    current_spa = None
    current_ent = None
    current_front = None
    current_message = None
    current_step = None
    isIncompleteMessage = False
    
    for linea in contenido_failures.splitlines():
    
    
        # SPA
        for spa in Test.ALLOWED_SPA:
            if spa in linea:
                current_spa = spa
                break
            
        # ENT
        for ent in Test.ALLOWED_ENT:
            if ent in linea:
                current_ent = ent
                break
            
        # FRONT
        if 'FRONT' in linea:
            match = re.search(r'(FRONT[^.]*)\.', linea)
            if match:
                current_front = match.group(1).strip()

        if isIncompleteMessage:
            # Elimina identificadores tipo [chrome #01-2]
            linea_limpia = re.sub(r'\[chrome\s*#\d+-\d+\]', '', linea)
        
            if 'Step' in linea_limpia:
                match_step = re.search(r'\(Step\s*\d+\)', linea_limpia)
                if match_step:
                    current_message += ' ' + linea_limpia[:match_step.start()].strip()
                    isIncompleteMessage = False
            else:
                current_message += ' ' + linea_limpia.strip()
        

        # Failed
        if 'Failed' in linea:
            match_failed = re.search(
                r'Failed:\s*(.+)(?:\s*\(Step\s*(\d+)\))?',
                linea
            )
            if match_failed:
                current_message = match_failed.group(1).strip()
                if 'Step' not in linea:
                    isIncompleteMessage = True

        
        if 'Step' in linea:
            match_step = re.search(r'\(Step\s*(\d+)\)', linea)
            if match_step: 
                current_step = match_step.group(1)

    
        # Cuando ya tenemos todo → crear Test
        if all([current_spa, current_ent, current_front, current_message, current_step]):
            current_message = re.sub(r'\(Step\s*\d+\).*', '', current_message).strip()
            listTest.append(Test(
                current_spa,
                current_ent,
                current_front,
                current_message,
                None,
                None,
                None,
                current_step,
                None,
                None
            ))
            contador += 1
    
            # Reset estado
            current_spa = None
            current_ent = None
            current_front = None
            current_message = None
            current_step = None
    
        
    print("La lista tiene un total de elementos de " + str(len(listTest)) + " tests")
    # 4.7 Completar campos para Excel
    Test.convertObjectForExcelLines(listTest, execution, user)

    # 4.8 Pasar a DataFrame
    df = tests_to_dataframe(listTest)

    # ORDENAR POR ALLURE MESSAGE
    df = df.sort_values(by="ALLURE MESSAGE", ascending=True)

    print(df)

    # 4.9 Ruta del Excel de salida
    ruta_excel = os.path.join(f"C:/Users/{user}/Downloads", f"resultados_#{execution}.xlsx")

    # 4.10 Guardar Excel
    df.to_excel(ruta_excel, index=False, engine="openpyxl")

    # 4.11 Formato (alineación) + Autofiltro en cabeceras + color cabeceras


    workbook = load_workbook(ruta_excel)
    sheet = workbook.active
    alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Aplicar alineación a todas las celdas
    for row in sheet.iter_rows(
        min_row=1,
        max_row=sheet.max_row,
        min_col=1,
        max_col=sheet.max_column,
    ):
        for cell in row:
            cell.alignment = alignment

    # Altura algo mayor para la cabecera
    sheet.row_dimensions[1].height = 22

    # 1) AJUSTAR ANCHO BASE DE TODAS LAS COLUMNAS (ej. 15)
    for col_idx in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = 15

    # 2) BUSCAR LA COLUMNA 'ALLURE MESSAGE' Y PONERLA A 25
    header_row = sheet[1]  # fila de cabeceras
    allure_col_index = None

    for cell in header_row:
        if cell.value == "ALLURE MESSAGE":
            allure_col_index = cell.column  # índice numérico de columna
            break

    if allure_col_index is not None:
        allure_col_letter = get_column_letter(allure_col_index)
        sheet.column_dimensions[allure_col_letter].width = 45  # ancho especial para ALLURE MESSAGE

    # FORMATO CABECERAS (fila 1): fondo azul oscuro y texto en blanco
    header_fill = PatternFill(
        start_color="000080",
        end_color="000080",
        fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:  # fila 1 completa (cabeceras)
        cell.fill = header_fill
        cell.font = header_font

    # ACTIVAR FILTROS EN LA FILA DE CABECERAS
    max_row = sheet.max_row
    max_col = sheet.max_column
    last_col_letter = get_column_letter(max_col)
    data_range = f"A1:{last_col_letter}{max_row}"
    sheet.auto_filter.ref = data_range

    workbook.save(ruta_excel)
    print(f"Excel generado en: {ruta_excel}")




if __name__ == "__main__":
    main()
